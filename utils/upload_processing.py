import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None


@dataclass
class ProcessResult:
    status: str
    summary: Dict[str, Any]
    preview: Dict[str, Any]
    errors: List[Dict[str, Any]]


SUPPORTED_EXTENSIONS = {
    ".csv": "csv",
    ".xlsx": "excel",
    ".xls": "excel",
    ".pdf": "pdf",
}


def _safe_preview_rows(df: pd.DataFrame, max_rows: int = 5) -> List[Dict[str, Any]]:
    if df.empty:
        return []
    head = df.head(max_rows)
    return head.to_dict(orient="records")


def detect_file_type(filename: str) -> Tuple[Optional[str], Optional[str]]:

    if not filename or "." not in filename:
        return None, None
    ext = "." + filename.rsplit(".", 1)[-1].lower()
    file_type = SUPPORTED_EXTENSIONS.get(ext)
    if not file_type:
        return ext, None
    return ext, file_type


def process_csv(data: bytes) -> ProcessResult:
    try:
        stream = io.BytesIO(data)
        # pandas can infer delimiter; let it default to comma.
        df = pd.read_csv(stream)
        columns = list(df.columns)
        summary = {"row_count": int(len(df)), "column_names": columns}
        preview = {"rows": _safe_preview_rows(df)}
        return ProcessResult(status="success", summary=summary, preview=preview, errors=[])
    except Exception as e:
        return ProcessResult(status="error", summary={}, preview={}, errors=[{"type": "csv_parse_error", "message": str(e)}])


def process_excel(data: bytes, *, all_sheets: bool = False) -> ProcessResult:
    if load_workbook is None:
        return ProcessResult(
            status="error",
            summary={},
            preview={},
            errors=[{"type": "missing_dependency", "message": "openpyxl is required for Excel parsing"}],
        )

    try:
        stream = io.BytesIO(data)
        wb = load_workbook(stream, read_only=True, data_only=True)
        sheets = wb.sheetnames

        def sheet_to_payload(sheet_name: str) -> Dict[str, Any]:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"sheet_name": sheet_name, "row_count": 0, "column_names": [], "preview": {"rows": []}}

            header = rows[0]
            header = [h for h in header]
            # Convert header to strings where possible
            column_names = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]

            data_rows = rows[1:]
            preview_rows = []
            for r in data_rows[:5]:
                rec = {}
                for i, v in enumerate(r[: len(column_names)]):
                    rec[column_names[i]] = v
                preview_rows.append(rec)

            return {
                "sheet_name": sheet_name,
                "row_count": max(len(data_rows), 0),
                "column_names": column_names,
                "preview": {"rows": preview_rows},
            }

        if all_sheets:
            sheet_payloads = [sheet_to_payload(s) for s in sheets]
            summary = {"sheets": sheets}
            preview = {"sheets": sheet_payloads}
        else:
            first = sheets[0] if sheets else ""
            sheet_payload = sheet_to_payload(first) if first else {}
            summary = {"sheets": sheets, "selected_sheet": first}
            preview = {"sheet": sheet_payload}

        return ProcessResult(status="success", summary=summary, preview=preview, errors=[])
    except Exception as e:
        return ProcessResult(status="error", summary={}, preview={}, errors=[{"type": "excel_parse_error", "message": str(e)}])


def process_pdf(data: bytes, *, max_pages_preview: int = 2, max_chars_per_page: int = 1200) -> ProcessResult:
    if PdfReader is None:
        return ProcessResult(
            status="error",
            summary={},
            preview={},
            errors=[{"type": "missing_dependency", "message": "pypdf is required for PDF parsing"}],
        )

    try:
        reader = PdfReader(io.BytesIO(data))
        page_count = len(reader.pages)
        pages_preview = []
        for i in range(min(page_count, max_pages_preview)):
            page = reader.pages[i]
            text = page.extract_text() or ""
            text = text.strip()[:max_chars_per_page]
            pages_preview.append({"page_index": i, "text": text})

        summary = {"page_count": int(page_count)}
        preview = {"pages": pages_preview}
        return ProcessResult(status="success", summary=summary, preview=preview, errors=[])
    except Exception as e:
        return ProcessResult(status="error", summary={}, preview={}, errors=[{"type": "pdf_parse_error", "message": str(e)}])


def process_by_type(file_type: str, data: bytes) -> ProcessResult:
    if file_type == "csv":
        return process_csv(data)
    if file_type == "excel":
        return process_excel(data)
    if file_type == "pdf":
        return process_pdf(data)
    return ProcessResult(status="error", summary={}, preview={}, errors=[{"type": "unsupported", "message": f"Unsupported file_type={file_type}"}])
